import subprocess
import time
import sys
import threading
import os, getopt
import os.path
from openpyxl import Workbook, load_workbook
import psutil
from nvitop import *
import math

def create_excel(file):
    wb = Workbook()
    ws = wb.active
    title = ["branch", "cpu_threads_num", "omp_num_threads", "rst_threads_num", "max_index", "grid_dim", "block_dim", "time (s)", 
             "thread_id", "parallel_num", "cmd", "error_flag", "start_time", "end_time",
             "gpu_sm_utilization", "gpu_memory", "cpu_percent", "host_memory"]
    ws.append(title)
    wb.save(file)

def write_excel(file, new_row):
    with lock:
        wb = load_workbook(file)
        ws = wb.active
        ws.append(new_row)
        wb.save(file)
        
def record_usage(pid, usage):
    while psutil.pid_exists(pid):
        try:
            devices = Device.all()  # or `Device.all()` to use NVML ordinal instead
            for device in devices:
                processes = device.processes()
                if len(processes) <= 0:
                    continue
                processes = GpuProcess.take_snapshots(processes.values(), failsafe=True)
                for p in processes:
                    if p.pid != pid:
                        continue
                    usage["gpu_sm_utilization"].append(p.gpu_sm_utilization)
                    usage["gpu_memory"].append(p.gpu_memory)
                    usage["cpu_percent"].append(HostProcess(pid).cpu_percent(interval=0.1))
                    usage["host_memory"].append(p.host_memory)
                    break
        except:
            break
    

def do_loop():
    rst = 0
    if branch == "openmp-depend-modify-rst":
        rst = rst_threads_num
    print("branch {}, parallel_num {}, cpu_threads_num {}, omp_num_threads {}, \
          rst_threads_num {}, max_index {}, loop_num {}, board size {}, grid_dim {}, block_dim {}".
          format(branch, parallel_num, cpu_threads_num, omp_num_threads, rst, max_index, loop_num, bd_size, grid_dim, block_dim))
    # file = data_txt.format(branch, parallel_num, cpu_threads_num, omp_num_threads)
    for j in range(loop_num):
        begin = time.time()
        begin_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(begin))
        cmd = ["./hybrid2", 
               "--cpu_threads_num={}".format(cpu_threads_num),
               "--max_count={}".format(10), 
               "--max_index={}".format(max_index),
               "--bd_size={}".format(bd_size), 
               "--num_moves={}".format(4),
               "--rst_threads_num={}".format(rst),
               "--grid_dim={}".format(grid_dim),
               "--block_dim={}".format(block_dim)]
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        # process = subprocess.Popen(["ls /bin"], shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        pid = process.pid
        
        usage = {"gpu_sm_utilization":[], "gpu_memory":[], "cpu_percent":[], "host_memory":[]}
        thread_record = threading.Thread(target=record_usage, args=(pid, usage, ))
        thread_record.start()
        
        res = process.communicate()
        end = time.time()
        end_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end))
        new_row = [branch, cpu_threads_num, omp_num_threads, rst, max_index, grid_dim, block_dim, (end - begin), pid, parallel_num, str(cmd)]
        if process.returncode != 0:
            new_row.append("error")
            print(res)
        else:
            new_row.append("ok")
        new_row.append(begin_str)
        new_row.append(end_str)
        print(new_row)
        # record actual num of thread
        thread_record.join()
        for key in usage:
            val = usage[key]
            val_str = str(val)
            new_row.append(val_str[1:len(val_str)-1])
        write_excel(file, new_row)

def run():
    # file = data_txt.format(branch, parallel_num, cpu_threads_num, omp_num_threads)
    if not os.path.isfile(file):
        create_excel(file)
    
    t = []
    for i in range(parallel_num):
        ti = threading.Thread(target=do_loop)
        t.append(ti)
        ti.start()

    for i in range(parallel_num):
        t[i].join()
        
    # print("sleep 5 min to seperate")
    # time.sleep(300)

def branch_run():
    if "pthread-norand" == branch:
        omp_num_threads = 0
        run()
    elif "openmp-depend" == branch:
        for omp_num_threads in omp_num_threads_list:
            os.environ["OMP_NUM_THREADS"] = str(omp_num_threads)
            run()
    elif "openmp-depend-modify-rst" == branch:
        for omp_num_threads in omp_num_threads_list:
            for rst_threads_num in rst_threads_num_list:
                os.environ["OMP_NUM_THREADS"] = "{}, {}".format(omp_num_threads, rst_threads_num)
                run()

# python auto_run.py --branch="pthread-norand" --loop_num=1 --parallel_num_list="1" 
# --cpu_threads_num_list="0" --omp_num_threads_list="2" --rst_threads_num_list="0" --board_size_list="15" 
# --max_index_range="10 20 10" --file=0626-test --grid_dim_range="128 2048 128" --block_dim_range="128 2048 128"
branch_list = []
loop_num = 20
parallel_num_list = []
cpu_threads_num_list = []
omp_num_threads_list = []
rst_threads_num_list = []
board_size_list = []
max_index_range = []
grid_dim_range = []
block_dim_range = []
gpu_dim_list = []
gpu_plus_cpu = []
file = "../data/test.xlsx"
options = "-b:-l:-p:-c:-o:-r:-s:-i:-f:"
long_options = ["branch=", "loop_num=", "parallel_num_list=", "cpu_threads_num_list=", "omp_num_threads_list=",
                "rst_threads_num_list=", "board_size_list=", "max_index_range=", "grid_dim_range=", "block_dim_range=", 
                "gpu_dim_list=", "gpu_plus_cpu=", "file="]
opts,args = getopt.getopt(sys.argv[1:], options, long_options)
for opt_name,opt_value in opts:
    if opt_name in ('-b', "--branch"):
        branch_list = opt_value.split(" ")
    if opt_name in ('-l', "--loop_num"):
        loop_num = int(opt_value)
    if opt_name in ('-p', "--parallel_num_list"):
        parallel_num_list = list(map(int, opt_value.split(" ")))
    if opt_name in ('-c', "--cpu_threads_num_list"):
        cpu_threads_num_list = list(map(int, opt_value.split(" ")))
    if opt_name in ('-o', "--omp_num_threads_list"):
        omp_num_threads_list = list(map(int, opt_value.split(" ")))
    if opt_name in ('-r', "--rst_threads_num_list"):
        rst_threads_num_list = list(map(int, opt_value.split(" ")))
    if opt_name in ('-s', "--board_size_list"):
        board_size_list = list(map(int, opt_value.split(" ")))
    if opt_name in ('-i', "--max_index_range"):
        max_index_range = list(map(int, opt_value.split(" ")))
    if opt_name in ("--grid_dim_range"):
        grid_dim_range = list(map(int, opt_value.split(" ")))
    if opt_name in ("--block_dim_range"):
        block_dim_range = list(map(int, opt_value.split(" ")))
    if opt_name in ("--gpu_dim_list"):
        gpu_dim_list = list(map(int, opt_value.split(" ")))
    if opt_name in ("--gpu_plus_cpu"):
        gpu_plus_cpu = list(map(int, opt_value.split(" ")))
    if opt_name in ('-f', "--file"):
        file = "../data/{}.xlsx".format(opt_value)

if len(gpu_dim_list) != 0 and len(gpu_plus_cpu) != 0:
    print("do not give both values: gpu_dim_list, gpu_plus_cpu")
    exit()
elif len(gpu_dim_list) != 0 and len(gpu_plus_cpu) == 0:
    temp = []
    for i in range(0, gpu_dim_list[len(gpu_dim_list)-2], 1):
        temp.append(math.pow(2, i))
    grid_dim_range = temp
    block_dim_range = temp
    
    temp = []
    for i in gpu_dim_list:
        temp.append(math.pow(2, i))
    gpu_dim_list = temp
else:
    grid_dim_range = range(grid_dim_range[0], grid_dim_range[1], grid_dim_range[2])
    block_dim_range = range(block_dim_range[0], block_dim_range[1], block_dim_range[2])
    
print(gpu_dim_list)
print("branch: {}, loop_num: {}, parallel_num_list: {}, cpu_threads_num_list: {}, omp_num_threads_list: {}, \
                rst_threads_num_list: {}, board_size_list: {}, max_index_range: {}, grid_dim_range: {}, block_dim_range: {}, file: {}".format(
                branch_list, loop_num, parallel_num_list, cpu_threads_num_list, omp_num_threads_list,
                rst_threads_num_list, board_size_list, max_index_range, grid_dim_range, block_dim_range, file))

lock = threading.Lock()
omp_num_threads = 0
rst_threads_num = 0
for branch in branch_list:
    os.system("git checkout " + branch)
    os.system("make")
    # data_txt = "../data/{}_{}_threads={}_pool={}.xlsx"

    for bd_size in board_size_list:
        for parallel_num in parallel_num_list:
            for cpu_threads_num in cpu_threads_num_list:
                for max_index in range(max_index_range[0], max_index_range[1], max_index_range[2]):
                    for block_dim in block_dim_range:
                    
                        if len(gpu_plus_cpu) != 0:
                            for gpc in gpu_plus_cpu:
                                grid_dim = int((gpc - cpu_threads_num*max_index) / block_dim)
                                if grid_dim <= 0: continue
                                run()
                                
                        if len(gpu_dim_list) != 0:
                            for grid_dim in grid_dim_range:
                                if (grid_dim * block_dim) not in gpu_dim_list:
                                    continue
                                run()
                            
